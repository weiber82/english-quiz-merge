from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseNotFound, JsonResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.http import FileResponse
from django.core.paginator import Paginator
from django.db.models import Count, Q, F
from .services.gpt_service import GPTExplanationService
from .services.openai_client import OpenAIClient
from .services.auth_service import AuthService
from .forms import QuestionForm, UserCreateForm
from .models import User, Favorite, Question, TestRecord, WrongQuestion, WeakTopic, Feedback, User, GptLog, ImproveSuggestion
from django.core.paginator import Paginator
from dotenv import load_dotenv
import json
import random
import os
import uuid
import openpyxl


load_dotenv()  # 讀取 .env 檔案

auth_service = AuthService()

def home(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')

    explanation = None
    if request.method == 'POST':
        question = request.POST.get('question')
        answer = request.POST.get('answer')

        import os
        client = OpenAIClient(api_key=os.getenv("OPENAI_API_KEY"))
        service = GPTExplanationService(gpt_client=client)
        explanation = service.explain(question, answer)

    return render(request, 'home.html', {'explanation': explanation})


def register_view(request):
    message = ""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        success, message = auth_service.register(username, password)
        if success:
            return redirect('login')  # 註冊成功就跳回登入頁
    return render(request, 'register.html', {'message': message})


def login_view(request):
    user_id = request.session.get('user_id')
    if user_id:
        return redirect('dashboard') 

    message = ""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        success, message = auth_service.login(request, username, password)
        if success:
            return redirect('dashboard')  # 登入成功導向主頁
    return render(request, 'login.html', {'message': message})


def logout_view(request):
    auth_service.logout(request)
    return redirect('login')  # 登出後導回首頁登入


def dashboard_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')

    return render(request, 'dashboard.html')


class TestConfig:
    def __init__(self, topic, count, include_gpt):
        self.topic = topic
        self.count = count
        self.include_gpt = include_gpt

    @classmethod
    def from_request(cls, request):
        return cls(
            topic=request.POST.get("topic"),
            count=int(request.POST.get("count")),
            include_gpt=request.POST.get("include_gpt")
        )
    

    def to_session_data(self):
        return {
            'test_config': {
                'topic': self.topic,
                'count': self.count,
                'include_gpt': self.include_gpt
            }
    }


class PracticeSession:
    def __init__(self, config):
        self.config = config
        self.test_result_id = str(uuid.uuid4())
        self.selected_questions = []

    # def initialize(self):
    #     self.selected_questions = random.sample(
    #         Question.get_by_topic(self.config.topic, self.config.include_gpt),
    #         k=self.config.count
    #     )

    #0603TEST

    def initialize(self):
        all_questions = Question.get_by_topic(self.config.topic, self.config.include_gpt)
        k = min(len(all_questions), self.config.count)  # ⭐ 保證不會抽超過題庫數
        self.selected_questions = random.sample(all_questions, k=k)
        

    def to_session_data(self):
        return {
            'test_result_id': self.test_result_id,
            'test_config': {
                'topic': self.config.topic,
                'count': self.config.count,
                'include_gpt': self.config.include_gpt,
            },
            'test_questions': [q.id for q in self.selected_questions],
            'answers': {}
        }


def start_test_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')

    if request.method == 'POST':
        # 清除舊紀錄
        request.session.pop('test_questions', None)
        request.session.pop('answers', None)

        config = TestConfig.from_request(request)
        session = PracticeSession(config)
        session.initialize()

        request.session.update(session.to_session_data())
        return redirect('test_question', question_index=0)

    return render(request, 'start_test.html')


def test_question_view(request, question_index):
    config = request.session.get('test_config')
    if not config:
        return redirect('start_test')

    if 'test_questions' not in request.session:
        topic = config['topic']
        count = config['count']
        include_gpt = config['include_gpt']

        all_questions = Question.get_by_topic(topic, include_gpt)
        selected = random.sample(all_questions, min(count, len(all_questions)))
        request.session['test_questions'] = [q.id for q in selected]
        request.session['answers'] = {}

    question_ids = request.session['test_questions']
    if question_index >= len(question_ids):
        return redirect('dashboard')

    selected_answer = None
    if request.method == 'POST':        
        selected_answer = request.POST.get('answer')
        question = Question.get_by_id(question_ids[question_index])

        answers = request.session.get('answers', {})
        answers[str(question.id)] = selected_answer
        request.session['answers'] = answers
        request.session.modified = True  # ⭐⭐ ←← 這行是關鍵！加上它才能真的寫進 session！

        user_id = request.session.get('user_id')
        test_result_id = request.session.get('test_result_id')

        print("✅ 寫入 session['answers']:", answers)

        if user_id and test_result_id:
            TestRecord.save_answer(user_id, question, selected_answer, test_result_id)

    else:
        question = Question.get_by_id(question_ids[question_index])

    return render(request, 'test_question.html', {
        'question': question,
        'index': question_index,
        'total': len(question_ids),
        'selected': selected_answer
    })


def test_result_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')

    test_result_id = request.session.get('test_result_id')
    if not test_result_id:
        return redirect('start_test')

    records = TestRecord.get_test_records(user_id, test_result_id)
    correct_count = TestRecord.get_correct_count(user_id, test_result_id)
    wrong_records = TestRecord.get_wrong_records(user_id, test_result_id)

    total = records.count()
    accuracy = round((correct_count / total) * 100, 2) if total else 0

    question_order = request.session.get('test_questions', [])

    indexed_wrong_records = []
    for record in wrong_records:
        try:
            seq = question_order.index(record.question.id) + 1
        except ValueError:
            seq = "?"
        is_starred = Favorite.is_starred(user_id, record.question.id)
        indexed_wrong_records.append({
            'record': record,
            'seq': seq,
            'is_starred': is_starred
        })

    # 移到最後頁面真正顯示完再清除
    response = render(request, 'test_result.html', {
        'correct_count': correct_count,
        'total': total,
        'accuracy': accuracy,
        'wrong_records': indexed_wrong_records
    })
    request.session.pop('test_result_id', None)
    return response


def gpt_detail_view(request):
    user_id = request.session.get('user_id')
    qid = int(request.GET.get('qid'))
    question = Question.get_by_id(qid)
    if not question:
        return HttpResponseNotFound("題目不存在")

    # 收藏狀態
    is_starred = Favorite.is_starred(user_id=user_id, question_id=question.id)

    # 回答記錄
    answers = request.session.get('answers', {})
    selected = answers.get(str(qid))

    # 預設標記
    next_index = None
    in_test_session = False
    just_finished_test = False
    from_test_flow = False

    test_questions = request.session.get('test_questions', [])

    try:
        idx = test_questions.index(str(qid))

        # 判斷是否為測驗流程中的正常跳轉
        if len(answers) == idx:
            from_test_flow = True
            in_test_session = True
            if idx + 1 < len(test_questions):
                next_index = idx + 1
            else:
                just_finished_test = True

        # 額外補一個 just_finished_test（例如從 result 再點詳解）
        elif idx + 1 == len(test_questions):
            just_finished_test = True

    except ValueError:
        pass

    # GPT 解釋
    client = OpenAIClient(api_key=os.getenv("OPENAI_API_KEY"))
    service = GPTExplanationService(gpt_client=client)
    explanation = service.explain(question, selected)

    return render(request, 'gpt_detail.html', {
        'question': question,
        'selected': selected,
        'explanation': explanation,
        'is_starred': is_starred,
        'next_index': next_index,
        'in_test_session': in_test_session,
        'just_finished_test': just_finished_test,
        'from_test_flow': from_test_flow,  # ✅ 多傳一個這個
    })


def user_management_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')

    current_user = User.get_by_id(user_id)
    if current_user.role != 'admin':
        return HttpResponseForbidden("你沒有權限瀏覽此頁面")

    if request.method == 'POST':
        target_id = request.POST.get('user_id')
        new_role = request.POST.get('role')

        if str(current_user.id) == target_id:
            messages.error(request, "無法修改自己的權限。")
            return redirect('user_management')

        target_user = User.get_by_id(target_id)
        target_user.role = new_role
        target_user.save()
        messages.success(request, f"使用者 {target_user.username} 已更新為 {new_role}。")
        return redirect('user_management')

    users = User.get_all()
    return render(request, 'user_management.html', {'users': users})


@csrf_exempt
def save_answer_view(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        qid = str(data.get('qid'))
        ans = data.get('answer')

        user_id = request.session.get('user_id')
        test_result_id = request.session.get('test_result_id')

        if user_id and test_result_id:
            question = Question.get_by_id(qid)
            TestRecord.save_answer(user_id, question, ans, test_result_id)

            #  如果答錯，就寫入 WrongQuestion
            if ans != question.answer:
                try:
                    current_user = User.get_by_id(user_id)
                    WrongQuestion.get_or_create(
                        user=current_user,
                        question=question,
                        defaults={'confirmed': False}
                    )
                except User.DoesNotExist:
                    print(f"❌ 無法寫入錯題，找不到 user_id={user_id} 的使用者")

        answers = request.session.get('answers', {})
        answers[qid] = ans
        request.session['answers'] = answers

        return JsonResponse({'status': 'ok'})

    return JsonResponse({'error': 'invalid request'}, status=400)
    

def update_note_view(request, fav_id):
    if request.method == 'POST':
        user_id = request.session.get('user_id')
        note_text = request.POST.get('note', '')
        Favorite.update_note(fav_id, user_id, note_text)  # 封裝在 model 內
        return redirect('favorite_questions')


def favorite_questions_view(request):
    user_id = request.session.get('user_id')
    favorites = Favorite.get_user_favorites(user_id)
    return render(request, 'favorite_questions.html', {'favorites': favorites})


@csrf_protect
def toggle_favorite_view(request, question_id):
    user_id = request.session.get('user_id')
    if not user_id:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'error': 'unauthorized'}, status=403)
        return redirect('login')

    is_starred = Favorite.toggle_star(user_id=user_id, question_id=question_id)

    # ✅ 如果是 AJAX 請求（例如 GPT 詳解頁），回傳 JSON，不刷新頁面
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'status': 'ok', 'starred': is_starred})

    # ✅ 否則仍支援表單 redirect（例如其他頁用 form POST）
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER')
    if next_url:
        return redirect(next_url)
    return redirect('favorite_questions')


def wrong_questions_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        # messages.error(request, "請先登入以查看錯題本。")
        return redirect('login')

    current_user = User.get_by_id(user_id)
    if not current_user:
        request.session.pop('user_id', None)
        # messages.error(request, "使用者資訊錯誤，請重新登入。")
        return redirect('login')

    selected_topic = request.GET.get('topic', None)
    # 獲取 view_mode 參數，預設為 'unfixed' (未學會)
    view_mode = request.GET.get('view_mode', 'unfixed') 

    page_title_base = "我的錯題本 (S5 功能)"

    if view_mode == 'fixed':
        # 呼叫獲取「已學會」錯題的方法
        wrong_list = WrongQuestion.get_fixed_by_user_and_topic(
            user=current_user, 
            topic=selected_topic
        )
        # 獲取「已學會」錯題的主題列表
        available_topics = WrongQuestion.get_distinct_topics_for_fixed_by_user(
            user=current_user
        )
        page_title = f"{page_title_base} - 已學會的題目"
    else:
        # 呼叫獲取「未學會」錯題的方法
        wrong_list = WrongQuestion.get_unconfirmed_by_user_and_topic(
            user=current_user, 
            topic=selected_topic
        )
        # 獲取「未學會」錯題的主題列表
        available_topics = WrongQuestion.get_distinct_topics_for_unconfirmed_by_user(
            user=current_user
        )
        page_title = f"{page_title_base} - 待複習的題目"
        view_mode = 'unfixed'

    context = {
        'wrong_questions': wrong_list,
        'available_topics': available_topics,
        'current_topic': selected_topic if selected_topic else 'all',
        'current_view_mode': view_mode,
        'page_title': page_title
    }
    return render(request, 'wrong_questions.html', context)


@require_POST # 標記為已學會是一個修改操作，用 POST
def mark_wrong_question_fixed_view(request, wrong_question_id):
    user_id_from_session = request.session.get('user_id')
    current_user = User.get_by_id(user_id_from_session)

    if not current_user:
        return JsonResponse({'status': 'error', 'message': '使用者未登入或 Session 已過期。'}, status=401)

    try:
        # 直接查詢，並確保是屬於當前使用者的錯題
        # wrong_question_entry = WrongQuestion.objects.get(id=wrong_question_id, user=current_user)  0603fix
        wrong_question_entry = WrongQuestion.get_by_id_and_user(wrong_question_id, current_user)

        # 呼叫在 WrongQuestion 模型中定義的 mark_as_learned (或類似名稱) 方法
        changed = wrong_question_entry.mark_as_learned() # 假設方法名是 mark_as_learned

        if changed:
            return JsonResponse({'status': 'success', 'message': '題目已成功標記為已學會。'})
        else:
            return JsonResponse({'status': 'info', 'message': '題目先前已被標記為已學會。'})

    except WrongQuestion.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': '找不到該錯題記錄或您沒有權限操作。'}, status=404)
    except Exception as e:
        # 實際部署時應使用 logging 模組
        print(f"Error in mark_wrong_question_fixed_view (wq_id: {wrong_question_id}): {type(e).__name__} - {e}")
        return JsonResponse({'status': 'error', 'message': '標記已學會時發生錯誤。'}, status=500)


def diagnose_weakness_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')

    current_user = User.get_by_id(user_id)
    if not current_user:
        request.session.pop('user_id', None)
        return redirect('login')

    analysis_result = {
        "weak_topics": [],
        "summary": "點擊「開始進行弱點分析」按鈕來查看您的 AI 診斷報告。"
    }

    if request.method == 'POST':
        all_user_wrong_questions_count = WrongQuestion.get_unconfirmed_by_user_and_topic(user=current_user).count()
        sample_count = all_user_wrong_questions_count // 2
        if sample_count == 0 and all_user_wrong_questions_count > 0:
            sample_count = 1

        wrong_questions_for_analysis = WrongQuestion.get_sample_for_weakness_analysis(current_user, sample_count)

        if wrong_questions_for_analysis:
            data_for_gpt = []
            for wq_object in wrong_questions_for_analysis:
                data_for_gpt.append({
                    'question_obj': wq_object.question,
                })

            import os
            from .services.gpt_service import GPTExplanationService
            from .services.openai_client import OpenAIClient

            if os.getenv("OPENAI_API_KEY"):
                client = OpenAIClient(api_key=os.getenv("OPENAI_API_KEY"))
                gpt_service = GPTExplanationService(gpt_client=client)

                predefined_topics = None
                current_analysis_from_service = gpt_service.analyze_weaknesses(
                    data_for_gpt, predefined_weak_topics=predefined_topics)

                analysis_result["summary"] = current_analysis_from_service.get("summary", "AI分析未能提供有效的文字摘要。")
                analysis_result["weak_topics"] = current_analysis_from_service.get("weak_topics", [])

                if analysis_result["weak_topics"]:
                    for topic_name in analysis_result["weak_topics"]:
                        if topic_name:
                            WeakTopic.update_or_create_weak_topic(current_user, topic_name)
            else:
                analysis_result["summary"] = "OpenAI API 金鑰未設定，無法進行 AI 分析。"
        else:
            analysis_result["summary"] = "沒有足夠的錯題進行分析（目前選取0題）。"

    existing_weak_topics = WeakTopic.get_weak_topics_for_user(current_user)

    context = {
        'analysis_summary': analysis_result.get("summary"),
        'existing_weak_topics': existing_weak_topics,
        'page_title': "AI 弱點診斷"
    }
    return render(request, 'weakness_analysis_result.html', context)


def grade_history_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')

    current_user = User.get_by_id(user_id)
    if not current_user:
        request.session.pop('user_id', None)
        return redirect('login')

    test_history_details = TestRecord.get_recent_test_session_summaries(user=current_user, limit=10)

    context = {
        'test_history': test_history_details,
        'page_title': "測驗歷史記錄"
    }
    return render(request, 'grade_history.html', context)

# A1 題庫管理 首頁
def manage_questions_index_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)
    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')
    
    selected_topic = request.GET.get('topic', None)

    current_user = User.get_by_id(user_id)
    if not current_user:
        request.session.pop('user_id', None)
        return redirect('login')

    question_list = Question.get_by_topic(
        topic=selected_topic,
    )
    
    if selected_topic == 'all' or not selected_topic:
        question_list = Question.get_all()
    else:
        question_list = Question.get_by_topic(
        topic=selected_topic,
        )
    for q in question_list:
        q.need_improve = ImproveSuggestion.need_improve_by_id(q.id)
    
    paginator = Paginator(question_list, 10)  # 每頁 10 題
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    

    topics = ['all', 'vocab', 'grammar', 'cloze', 'reading']

    context = {
        'all_questions': question_list,
        'all_questions': page_obj,
        'topics': topics,
        'current_topic': selected_topic if selected_topic else 'all'
    }
    return render(request, 'manage_questions/index.html', context)

# A1 題庫管理 新增
def manage_questions_create_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)
    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('manage_questions_index')  # 你可以換成你要導向的頁面
    else:
        form = QuestionForm()

    return render(request, 'manage_questions/create.html', {'form': form})
    context = {
        'form': form
    }

    return render(request, 'manage_questions/create.html', context)

# A1 題庫管理 編輯
def manage_questions_edit_view(request, question_id):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)
    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')
    
    current_user = User.get_by_id(user_id)
    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')

    question = Question.get_by_id(question_id)

    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            return redirect('manage_questions_index')
    else:
        form = QuestionForm(instance=question)

    return render(request, 'manage_questions/edit.html', {'form': form, 'question': question})
    
    context = {
        'form': form,
        'question': question
    }

    return render(request, 'manage_questions/edit.html', context)

# A1 題庫管理 刪除
def manage_questions_delete_view(request, question_id):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)
    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')
    
    success = Question.get_by_id(question_id)

    if success:
        Question.delete_by_id(question_id)
        messages.success(request, "題目已成功刪除。") 
    else:
        messages.warning(request, "這筆資料不存在或已被刪除。")
    
    return redirect('manage_questions_index')

# A1 題庫管理 已解決須檢討題目
def manage_questions_edit_view(request, question_id):
    question = Question.get_by_id(question_id)
    form = QuestionForm(instance=question)

    suggestions = ImproveSuggestion.list_by_question(question_id)

    context = {
        'form': form,
        'question': question,
        'suggestions': suggestions,
    }
    return render(request, 'manage_questions/edit.html', context)


@require_POST
def resolve_suggestion(request):
    suggestion_id = request.POST.get('suggestion_id')
    print(suggestion_id)
    if not suggestion_id:
        return JsonResponse({'success': False, 'message': '缺少 suggestion_id'})

    success = ImproveSuggestion.mark_resolved_by_id(suggestion_id)
    return JsonResponse({'success': success})

# A2 Excel題庫匯入 首頁
def import_excel_index_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)
    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')

    preview_data = request.session.get('preview_questions', [])
    paginator = Paginator(preview_data, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    error = request.session.pop('import_error', None)

    context = {
        'preview_questions': page_obj,
        'error': error
    }
    return render(request, 'import_excel/index.html', context)

# A2 Excel題庫匯入 下載檔案模板
def import_excel_download_template_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)
    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')

    file_path = os.path.join(settings.BASE_DIR, 'static', 'templates', 'template.xlsx')
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='template.xlsx')

# A2 Excel題庫匯入 上傳檔案
def import_excel_upload_file_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)
    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')

    if request.method == 'POST' and request.FILES.getlist('excel_file'):
        preview_data = request.session.get('preview_questions', [])

        uploaded_files = request.FILES.getlist('excel_file')

        expected = ['主題', '題目內容', '選項A', '選項B', '選項C', '選項D', '正確答案']

        for file in uploaded_files:
            try:
                wb = openpyxl.load_workbook(file)
                ws = wb.active

                header = [cell.value for cell in ws[1]]
                if header != expected:
                    request.session['import_error'] = f'{file.name} 欄位格式錯誤，請使用模板上傳。'
                    return redirect('import_excel_index')

                for row in ws.iter_rows(min_row=2, values_only=True):
                    topic, content, a, b, c, d, answer = row
                    preview_data.append({
                        'topic': topic,
                        'content': content,
                        'options': {'A': a, 'B': b, 'C': c, 'D': d},
                        'answer': answer,
                        'created_dt': '尚未儲存'
                    })
            except Exception as e:
                request.session['import_error'] = f'{file.name} 檔案讀取錯誤：{str(e)}'
                return redirect('import_excel_index')

        # 更新 session
        request.session['preview_questions'] = preview_data
        return redirect('import_excel_index')

    else:
        request.session['import_error'] = '未收到檔案。'
        return redirect('import_excel_index')

# A2 Excel題庫匯入 取消預覽
def import_excel_cancel_preview_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)
    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')

    request.session.pop('preview_questions', None)
    request.session['import_error'] = '已取消預覽匯入資料。'
    return redirect('import_excel_index')

# A2 Excel題庫匯入 匯入題庫
def import_excel_confirm_save_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)
    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')

    preview_data = request.session.get('preview_questions', [])

    if not preview_data:
        request.session['import_error'] = '找不到預覽資料，請重新上傳 Excel。'
        return redirect('import_excel_index')

    # 將題目存入資料庫
    for q in preview_data:
        try:
            # print(q)
            Question.create_from_excel(q)
        except Exception as e:
            print('寫入失敗:', e)

    # 清除 session，避免重複匯入
    request.session.pop('preview_questions', None)

    return redirect('import_excel_index')


# A3 使用者權限管理 首頁
def manage_users_index_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)
    # print(current_user.role)
    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')

    users = User.get_all()

    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'users': page_obj
    }

    return render(request, 'manage_users/index.html', context)

# A3 使用者權限管理 新增
def manage_users_create_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)

    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')

    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "使用者新增成功")
            return redirect('manage_users_index')
    else:
        form = UserCreateForm()
    
    context = {
        'form': form
    }
    return render(request, 'manage_users/create.html', context)

# A3 使用者權限管理 編輯
def manage_users_edit_view(request, u_id):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)

    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')

    user_obj = User.get_by_id(u_id)
    if not user_obj:
        messages.warning(request, "找不到該使用者，無法編輯。")
        return redirect('manage_users_index')

    if request.method == 'POST':
        form = UserCreateForm(request.POST, instance=user_obj)
        if form.is_valid():
            form.save()
            messages.success(request, '使用者更新成功。')
            return redirect('manage_users_index')
    else:
        form = UserCreateForm(instance=user_obj)

    return render(request, 'manage_users/edit.html', {
        'form': form,
        'user_obj': user_obj,
    })

# A3 使用者權限管理 刪除
def manage_users_delete_view(request, u_id):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)

    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')
    
    if int(user_id) == int(u_id):
        messages.warning(request, "無法刪除自己的帳號。")
        return redirect('manage_users_index')

    success = User.delete_by_id(u_id)
    print("刪除回傳：", success)

    if success:
        messages.success(request, '使用者已刪除。')
    else:
        messages.warning(request, '找不到該使用者，無法刪除。')

    return redirect('manage_users_index')


# A4 測驗分析與報表下載 首頁
def analysis_index_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)

    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')
    
    if request.method == 'GET' and 'topic' not in request.GET and 'qid' not in request.GET:
        topic_dist = Question.get_topic_distribution()
        topic_acc = TestRecord.get_topic_accuracy()

        accuracy_map = {item['topic']: item['accuracy'] for item in topic_acc}

        for item in topic_dist:
            topic = item['topic']
            item['accuracy'] = accuracy_map.get(topic, 0)

        context = {
            'topic_distribution': topic_dist,
        }
        return render(request, 'analysis/index.html', context)
    if 'topic' in request.GET:
        topic = request.GET.get('topic')
        if topic == 'all':
            questions = Question.get_all()
        else:
            questions = Question.get_by_topic(topic)
        topics = ['all', 'vocab', 'grammar', 'cloze', 'reading']
        
        paginator = Paginator(questions, 10)  # 每頁 10 題
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'questions': page_obj,
            'topics': topics,
            'current_topic': topic
        }

        return render(request, 'analysis/topic_question_list.html', context)

    # --- Case C: 題目選項統計圖 (AJAX) ---
    if 'qid' in request.GET:
        qid = request.GET.get('qid')
        question = Question.get_by_id(qid)
        stats = TestRecord.get_question_stats(qid)

        response_data = {
            'stats': stats['stats'],
            'ratio': stats['ratio'],
            'pass_rate': stats['pass_rate'],
            'total': stats['total'],
            'answer': question.answer if question else None
        }

        return JsonResponse(response_data)


# A4 測驗分析與報表下載 發送建議
@require_POST
def analysis_submit_improve_suggestion(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return JsonResponse({'success': False, 'message': '尚未登入'}, status=403)

    question_id = request.POST.get('question_id')
    suggestion_text = request.POST.get('suggestion', '').strip()

    if not question_id or not suggestion_text:
        return JsonResponse({'success': False, 'message': '缺少欄位'}, status=400)

    try:
        ImproveSuggestion.add_suggestion(
            question_id=question_id,
            user_id=user_id,
            suggestion=suggestion_text
        )
        return JsonResponse({'success': True, 'message': '建議已儲存'})
    except ImproveSuggestion.DoesNotExist:
        return JsonResponse({'success': False, 'message': '找不到題目'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

# A5 題目品質查詢與回覆 首頁
def quality_index_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    current_user = User.get_by_id(user_id)

    if not current_user or current_user.role != 'admin':
        messages.warning(request, "您沒有權限瀏覽此頁面。")
        return redirect('dashboard')
    
    if request.method == 'GET' and 'topic' not in request.GET and 'qid' not in request.GET:
        topic_dist = Question.get_topic_distribution()
        topic_rating = Feedback.get_average_rating_by_topic()
        rating_map = {item['topic']: item['avg_rating'] for item in topic_rating}

        for item in topic_dist:
            topic = item['topic']
            item['avg_rating'] = round(float(rating_map.get(topic, 0) or 0), 2)

        context = {
            'topic_distribution': topic_dist,
        }
        return render(request, 'quality/index.html', context)
    if 'topic' in request.GET:
        topic = request.GET.get('topic')
        if topic == 'all':
            questions = Question.get_all()
        else:
            questions = Question.get_by_topic(topic)
        
        for q in questions:
            q.avg_rating = Feedback.get_average_rating_for_question_id(q.id)
            q.feedbacks = Feedback.get_feedbacks_for_question(q)
            stats = TestRecord.get_question_stats(q.id)
            q.pass_rate = stats['pass_rate']
            print(f"Q{q.id} stats: {stats['total']}")
        
        topics = ['all', 'vocab', 'grammar', 'cloze', 'reading']
        
        paginator = Paginator(questions, 10)  # 每頁 10 題
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'questions': page_obj,
            'topics': topics,
            'current_topic': topic
        }

        return render(request, 'quality/topic_question_list.html', context)


# A5 題目品質查詢與回覆 發送建議
@require_POST
def quality_submit_improve_suggestion(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return JsonResponse({'success': False, 'message': '尚未登入'}, status=403)

    question_id = request.POST.get('question_id')
    suggestion_text = request.POST.get('suggestion', '').strip()

    if not question_id or not suggestion_text:
        return JsonResponse({'success': False, 'message': '缺少欄位'}, status=400)

    try:
        ImproveSuggestion.add_suggestion(
            question_id=question_id,
            user_id=user_id,
            suggestion=suggestion_text
        )
        return JsonResponse({'success': True, 'message': '建議已儲存'})
    except ImproveSuggestion.DoesNotExist:
        return JsonResponse({'success': False, 'message': '找不到題目'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

class WrongChallengeSession:
    def __init__(self, user):
        self.user = user

class WrongChallengeSession:
    def __init__(self, user):
        self.user = user
        self.selected_questions = []

    def initialize(self, count=5):
        wrong_questions = WrongQuestion.get_unfixed_by_user(self.user)
        self.selected_questions = random.sample(
            [wq.question for wq in wrong_questions],
            min(count, len(wrong_questions))
        )

def start_wrong_challenge(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')

    current_user = User.get_by_id(user_id)
    session = WrongChallengeSession(current_user)
    session.initialize(count=5)

    request.session['challenge_questions'] = [q.id for q in session.selected_questions]
    request.session.modified = True

    return render(request, 'wrong_challenge.html', {
        'questions': session.selected_questions
    })

def submit_wrong_challenge(request):
    user_id = request.session.get('user_id')
    if not user_id or request.method != 'POST':
        return redirect('login')

    current_user = User.get_by_id(user_id)
    question_ids = request.session.get('challenge_questions', [])
    correct_count = 0

    for qid in question_ids:
        selected = request.POST.get(f'question_{qid}')
        q = Question.get_by_id(qid)
        if selected == q.answer:
            correct_count += 1
        else:
            WrongQuestion.mark_as_wrong(current_user, q)

    score = round((correct_count / len(question_ids)) * 100) if question_ids else 0

    return render(request, 'challenge_result.html', {
        'score': score,
        'total': len(question_ids),
        'correct': correct_count
    })


# --- S8 新增的視圖函式 ---
@require_POST 
def submit_feedback_view(request):
    try:
        # 判斷請求內容類型
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            question_id_str = data.get('question_id')
            rating_str = data.get('rating')
            comment = data.get('comment', '')
        else:
            question_id_str = request.POST.get('question_id')
            rating_str = request.POST.get('rating')
            comment = request.POST.get('comment', '')

        # 基本的資料驗證
        if not question_id_str or not rating_str:
            return JsonResponse({'status': 'error', 'message': '缺少題目ID或評分。'}, status=400)

        try:
            question_id = int(question_id_str)
            rating = int(rating_str)
        except ValueError:
            return JsonResponse({'status': 'error', 'message': '題目ID或評分必須是有效的數字。'}, status=400)

        # --- 使用模型方法進行資料存取 ---
        user_id_from_session = request.session.get('user_id')
        if not user_id_from_session:
            return JsonResponse({'status': 'error', 'message': '使用者未登入或 Session 已過期。'}, status=401)
        
        # 透過 User 模型的類別方法獲取使用者實例
        user_instance = User.get_by_id(user_id_from_session)
        if not user_instance:
            return JsonResponse({'status': 'error', 'message': '找不到使用者資訊 (ID: {}).'.format(user_id_from_session)}, status=404)

        # 透過 Question 模型的類別方法獲取題目實例
        question_instance = Question.get_by_id(question_id)
        if not question_instance:
            return JsonResponse({'status': 'error', 'message': '找不到指定的題目 (ID: {}).'.format(question_id)}, status=404)
        
        # 呼叫 Feedback 模型的類別方法來新增或更新評價
        feedback_obj, created = Feedback.add_or_update_feedback(
            user_instance=user_instance,
            question_instance=question_instance,
            rating_value=rating,
            comment_text=comment
        )
        # --- 資料存取結束 ---

        message = '評價已成功更新！' if not created else '評價已成功提交！'
        # 可以回傳 feedback_obj 的更多資訊，如果前端需要的話
        return JsonResponse({
            'status': 'success', 
            'message': message, 
            'feedback_id': feedback_obj.id,
            'created': created 
        })

    except ValueError as ve: 
        return JsonResponse({'status': 'error', 'message': str(ve)}, status=400)
    except json.JSONDecodeError: 
        return JsonResponse({'status': 'error', 'message': '無效的 JSON 格式。'}, status=400)
    except Exception as e:
        print(f"內部錯誤 in submit_feedback_view: {type(e).__name__} - {e}") # 暫時用 print 偵錯
        return JsonResponse({'status': 'error', 'message': '提交評價時發生未預期的伺服器錯誤。'}, status=500)
    
    
def my_feedback_view(request):
    user_id_from_session = request.session.get('user_id')
    current_user = User.get_by_id(user_id_from_session)

    if not current_user:
        messages.error(request, "請先登入以查看您的評價。")
        return redirect('login')

    # 獲取前端選擇的篩選主題 (從 GET 請求參數)
    selected_topic = request.GET.get('topic', None)

    # 根據選擇的主題篩選評價
    if selected_topic and selected_topic.lower() == 'all':
        user_feedbacks = Feedback.get_feedbacks_by_user(user_instance=current_user)
    else:
        user_feedbacks = Feedback.get_feedbacks_by_user(user_instance=current_user, topic_name=selected_topic)
    
    # 獲取該使用者評價過的所有不重複主題，用於生成篩選按鈕
    available_topics = Feedback.get_distinct_topics_for_user_feedback(user_instance=current_user)

    context = {
        'feedbacks': user_feedbacks,
        'page_title': '我提交的評價',
        'available_topics': available_topics, 
        'current_topic': selected_topic if selected_topic else 'all' 
    }
    return render(request, 'my_feedback.html', context)


@require_POST 
def update_feedback_view(request, feedback_id): # 
    try:
        user_id_from_session = request.session.get('user_id')
        current_user = User.get_by_id(user_id_from_session)
        if not current_user:
            return JsonResponse({'status': 'error', 'message': '使用者未登入或 Session 已過期。'}, status=401)

        # 使用 Feedback 模型的方法獲取屬於該使用者的特定評價記錄
        feedback_to_update = Feedback.get_feedback_by_id_and_user(feedback_id, current_user)

        if not feedback_to_update:
            return JsonResponse({'status': 'error', 'message': '找不到該評價記錄或您沒有權限編輯此評價。'}, status=404)

        # 從請求中獲取新的評分和評論
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            new_rating_str = data.get('rating')
            new_comment = data.get('comment', feedback_to_update.comment if feedback_to_update.comment is not None else "") 
        else:
            new_rating_str = request.POST.get('rating')
            new_comment = request.POST.get('comment', feedback_to_update.comment if feedback_to_update.comment is not None else "")

        if not new_rating_str:
            return JsonResponse({'status': 'error', 'message': '缺少評分資訊。'}, status=400)
        
        try:
            new_rating = int(new_rating_str)
        except ValueError:
            return JsonResponse({'status': 'error', 'message': '評分必須是有效的數字。'}, status=400)

        # 呼叫 Feedback 實例的 update_details 方法來更新評價
        feedback_to_update.update_details(new_rating=new_rating, new_comment=new_comment)
        
        # 準備回傳給前端的更新後 feedback 資料
        updated_feedback_data = {
            'id': feedback_to_update.id,
            'rating': feedback_to_update.rating,
            'comment': feedback_to_update.comment,
            'question_content': feedback_to_update.question.content[:30] + "..." if feedback_to_update.question else "N/A",
            'updated_at': feedback_to_update.updated_at.strftime("%Y-%m-%d %H:%M")
        }
        
        return JsonResponse({
            'status': 'success', 
            'message': '評價已成功更新！',
            'feedback': updated_feedback_data
        })

    except ValueError as ve: 
        return JsonResponse({'status': 'error', 'message': str(ve)}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': '無效的 JSON 格式。'}, status=400)
    except Exception as e:
        print(f"內部錯誤 in update_feedback_view (feedback_id: {feedback_id}): {type(e).__name__} - {e}") #用 print 偵錯
        return JsonResponse({'status': 'error', 'message': '更新評價時發生未預期的伺服器錯誤。'}, status=500)


def generate_similar_question_page_view(request):
    user_id_from_session = request.session.get('user_id')
    current_user = User.get_by_id(user_id_from_session)

    if not current_user:
        return render(request, 'recommend_question.html', {
            'error': '請先登入後再使用此功能（Session 遺失）。'
        })

    user = current_user

    wrong_q_list = WrongQuestion.get_unfixed_by_user(user)
    if not wrong_q_list:
        return render(request, 'recommend_question.html', {
            'error': '目前無可推薦的錯題，請先完成更多測驗。'
        })

    selected_wrong = random.choice(list(wrong_q_list))
    original_q = selected_wrong.question

    # ✅ 用你自訂的方法取最新紀錄
    last_record = TestRecord.get_latest_by_user_and_question(user, original_q)

    if not last_record or last_record.is_correct:
        return render(request, 'recommend_question.html', {
            'error': '找不到該題的錯誤選項。'
        })

    wrong_option = last_record.selected_option

    log_entry = GptLog.get_cached_similar_question(original_q, wrong_option)
    if log_entry:
        q = log_entry.generated_question
        source = 'GPT快取（Cache）'
        explanation = log_entry.explanation or ''
    else:
        api_key = os.getenv("OPENAI_API_KEY")
        gpt_client = OpenAIClient(api_key)
        gpt_service = GPTExplanationService(gpt_client=gpt_client)

        gpt_result = gpt_service.generate_similar_question(original_q, wrong_option)

        if not gpt_result:
            return render(request, 'recommend_question.html', {
                'error': 'GPT 產生失敗或格式錯誤，請稍後再試。'
            })

        q = Question.create_from_gpt(
            content=gpt_result['content'],
            options=gpt_result['options'],
            answer=gpt_result['answer'],
            topic=original_q.topic
        )
        GptLog.create_log_entry(original_q, q, wrong_option, explanation=gpt_result.get('explanation', ''))
        source = 'GPT 生成'
        explanation = gpt_result.get('explanation', '')

    return render(request, 'recommend_question.html', {
        'question': q,
        'source': source,
        'explanation': explanation
    })
